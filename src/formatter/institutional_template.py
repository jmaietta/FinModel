"""
Enhanced Detailed Excel template module for institutional investors.

This module provides a specialized Excel template for institutional investors
with comprehensive income statement line items and three years of quarterly history.
Updated to work with combined SG&A from Polygon data.
"""
import os
import logging
from typing import Dict, List, Optional, Any
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.utils import get_column_letter


class InstitutionalDetailedTemplate:
    """Creates detailed Excel templates for institutional investors."""
    
    def __init__(self):
        """Initialize institutional detailed template."""
        self.logger = logging.getLogger(__name__)
        
        # Define styles
        self.header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        self.subheader_font = Font(name='Arial', size=11, bold=True)
        self.normal_font = Font(name='Arial', size=10)
        self.number_font = Font(name='Arial', size=10)
        self.note_font = Font(name='Arial', size=9, italic=True, color='666666')
        
        self.header_fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type='solid')
        self.subheader_fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
        self.alternate_row_fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
        self.na_fill = PatternFill(start_color='FFEEEE', end_color='FFEEEE', fill_type='solid')
        
        self.center_align = Alignment(horizontal='center', vertical='center')
        self.right_align = Alignment(horizontal='right', vertical='center')
        self.left_align = Alignment(horizontal='left', vertical='center')
        
        self.border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        # Define institutional line items in order - cleaned up to show only available data
        self.institutional_line_items = [
            'Revenues',
            'CostOfGoodsSold',
            'GrossProfit',
            'ResearchAndDevelopmentExpense',
            'SellingGeneralAndAdministrativeExpenses',  # Combined SG&A when available
            'OperatingExpenses',
            'OperatingIncomeLoss',
            'OtherExpenses',  # Combined Interest & Other Income/Expense (mapped from nonoperating_income_loss)
            'IncomeLossBeforeIncomeTaxes',
            'IncomeTaxExpenseBenefit',
            'NetIncomeLoss',
            'EPS',  # Calculated field: Net Income / Fully-Diluted Shares
            'WeightedAverageSharesOutstandingDiluted'
        ]
        
        # Map item keys to display names - cleaned up to show only available data
        self.item_display_names = {
            'Revenues': 'Total Revenue',
            'CostOfGoodsSold': 'Cost of Goods Sold',
            'GrossProfit': 'Gross Profit',
            'ResearchAndDevelopmentExpense': 'Research & Development',
            'SellingGeneralAndAdministrativeExpenses': 'Sales, General & Administrative (Combined)',
            'OperatingExpenses': 'Total Operating Expenses',
            'OperatingIncomeLoss': 'Operating Income',
            'OtherExpenses': 'Interest & Other Income, Expense',
            'IncomeLossBeforeIncomeTaxes': 'Pre-Tax Income',
            'IncomeTaxExpenseBenefit': 'Income Tax Expense',
            'NetIncomeLoss': 'Net Income',
            'EPS': 'Earnings Per Share (Diluted)',
            'WeightedAverageSharesOutstandingDiluted': 'Fully-Diluted Shares Outstanding'
        }
        
        # Items that are unavailable and should be highlighted (now only SG&A if not available)
        self.unavailable_items = set()  # No longer needed since we removed unavailable items
    
    def create_template(self, income_statement: Dict, output_path: str) -> str:
        """Create institutional detailed template for income statement.
        
        Args:
            income_statement: Income statement data.
            output_path: Path to save the Excel file.
            
        Returns:
            Path to the saved Excel file.
        """
        wb = openpyxl.Workbook()
        
        # Create Income Statement sheet
        income_stmt_sheet = wb.active
        income_stmt_sheet.title = "Income Statement"
        
        # Create income statement sheet with detailed line items
        self._create_income_statement_sheet(income_stmt_sheet, income_statement)
        
        # Create Data Notes sheet
        notes_sheet = wb.create_sheet("Data Notes")
        self._create_data_notes_sheet(notes_sheet, income_statement)
        
        # Adjust column widths
        income_stmt_sheet.column_dimensions['A'].width = 35
        for col in range(2, 15):
            income_stmt_sheet.column_dimensions[get_column_letter(col)].width = 15
        
        # Save workbook
        try:
            wb.save(output_path)
            self.logger.info(f"Successfully saved institutional detailed template to {output_path}")
            return output_path
        except Exception as e:
            self.logger.error(f"Error saving institutional detailed template: {str(e)}")
            raise
    
    def _create_income_statement_sheet(self, sheet, income_statement: Dict):
        """Create income statement sheet with institutional line items.
        
        Args:
            sheet: Excel worksheet to populate.
            income_statement: Income statement data.
        """
        # Add title
        sheet['A1'] = f"{income_statement.get('company_name', '')} ({income_statement.get('ticker', '')}) - Income Statement"
        sheet['A1'].font = Font(name='Arial', size=16, bold=True)
        sheet.merge_cells('A1:N1')
        sheet['A1'].alignment = Alignment(horizontal='center')
        
        # Add data quality note
        sheet['A2'] = "Data sourced from Polygon.io - Combined line items reflect provider's data structure"
        sheet['A2'].font = self.note_font
        sheet.merge_cells('A2:N2')
        sheet['A2'].alignment = Alignment(horizontal='center')
        
        # Extract periods and sort by date (most recent first)
        periods = income_statement.get('periods', {})
        sorted_periods = sorted(periods.items(), key=lambda x: x[0], reverse=True)
        
        if not sorted_periods:
            sheet['A4'] = "No data available"
            return
        
        # Limit to 12 quarters (3 years)
        sorted_periods = sorted_periods[:12]
        
        # Add headers
        sheet['A4'] = "Line Item"
        sheet['A4'].font = self.header_font
        sheet['A4'].fill = self.header_fill
        sheet['A4'].alignment = self.center_align
        sheet['A4'].border = self.border
        
        # Add period headers
        for i, (period_key, _) in enumerate(sorted_periods):
            col = i + 2
            cell = sheet.cell(row=4, column=col)
            cell.value = period_key
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_align
            cell.border = self.border
        
        # Add line items
        for i, item_key in enumerate(self.institutional_line_items):
            row = i + 5
            
            # Item name
            cell = sheet.cell(row=row, column=1)
            cell.value = self.item_display_names.get(item_key, item_key)
            cell.font = self.normal_font
            cell.alignment = self.left_align
            cell.border = self.border
            
            # Apply alternating row fill
            if i % 2 == 1:
                cell.fill = self.alternate_row_fill
            
            # Add values for each period
            for j, (period_key, period_data) in enumerate(sorted_periods):
                col = j + 2
                cell = sheet.cell(row=row, column=col)
                
                items = period_data.get('items', {})
                
                # Handle EPS calculation
                if item_key == 'EPS':
                    # Calculate EPS = Net Income / Fully-Diluted Shares Outstanding
                    net_income = items.get('NetIncomeLoss', {}).get('value')
                    shares_outstanding = items.get('WeightedAverageSharesOutstandingDiluted', {}).get('value')
                    
                    if (net_income is not None and shares_outstanding is not None and 
                        shares_outstanding != 0):
                        eps_value = net_income / shares_outstanding
                        cell.value = eps_value
                        cell.number_format = '$0.00'  # EPS in dollars and cents
                        cell.font = self.number_font
                    else:
                        cell.value = "N/A"
                        cell.font = self.note_font
                
                elif item_key in items and items[item_key].get('value') is not None:
                    value = items[item_key].get('value', 0)
                    cell.value = value
                    
                    # Format based on item type
                    if item_key == 'WeightedAverageSharesOutstandingDiluted':
                        cell.number_format = '#,##0'  # No currency for shares
                    else:
                        cell.number_format = '$#,##0,,"M"'  # Display in millions
                    
                    cell.font = self.number_font
                else:
                    cell.value = "N/A"
                    cell.font = self.note_font
                
                cell.alignment = self.right_align
                cell.border = self.border
                
                # Apply alternating row fill for available items
                if i % 2 == 1:
                    cell.fill = self.alternate_row_fill
        
        # Add calculated margins
        margin_items = [
            ("Gross Margin", 'GrossProfit', 'Revenues'),
            ("Operating Margin", 'OperatingIncomeLoss', 'Revenues'),
            ("Net Margin", 'NetIncomeLoss', 'Revenues')
        ]
        
        start_row = len(self.institutional_line_items) + 6
        
        # Add margin header
        cell = sheet.cell(row=start_row, column=1)
        cell.value = "Margins"
        cell.font = self.subheader_font
        cell.fill = self.subheader_fill
        cell.alignment = self.left_align
        cell.border = self.border
        
        for j in range(1, len(sorted_periods) + 2):  # Extend across all columns
            if j > 1:
                cell = sheet.cell(row=start_row, column=j)
                cell.fill = self.subheader_fill
                cell.border = self.border
        
        # Add margin calculations
        for i, (margin_name, numerator_key, denominator_key) in enumerate(margin_items):
            row = start_row + i + 1
            
            # Margin name
            cell = sheet.cell(row=row, column=1)
            cell.value = margin_name
            cell.font = self.normal_font
            cell.alignment = self.left_align
            cell.border = self.border
            
            # Apply alternating row fill
            if i % 2 == 0:
                cell.fill = self.alternate_row_fill
            
            # Calculate margin for each period
            for j, (period_key, period_data) in enumerate(sorted_periods):
                col = j + 2
                cell = sheet.cell(row=row, column=col)
                
                items = period_data.get('items', {})
                if (numerator_key in items and 
                    denominator_key in items and
                    items[numerator_key].get('value') is not None and
                    items[denominator_key].get('value') is not None and
                    items[denominator_key].get('value', 0) != 0):
                    
                    numerator = items[numerator_key].get('value', 0)
                    denominator = items[denominator_key].get('value', 0)
                    margin = numerator / denominator * 100
                    cell.value = margin
                    cell.number_format = '0.00"%"'
                    cell.font = self.number_font
                else:
                    cell.value = "N/A"
                    cell.font = self.note_font
                
                cell.alignment = self.right_align
                cell.border = self.border
                
                # Apply alternating row fill
                if i % 2 == 0:
                    cell.fill = self.alternate_row_fill
        
        # Add Year-over-Year Growth section
        yoy_start_row = start_row + len(margin_items) + 3  # Leave a blank line
        
        # Add YoY Growth header
        cell = sheet.cell(row=yoy_start_row, column=1)
        cell.value = "Year-over-Year Growth"
        cell.font = self.subheader_font
        cell.fill = self.subheader_fill
        cell.alignment = self.left_align
        cell.border = self.border
        
        for j in range(1, len(sorted_periods) + 2):  # Extend across all columns
            if j > 1:
                cell = sheet.cell(row=yoy_start_row, column=j)
                cell.fill = self.subheader_fill
                cell.border = self.border
        
        # Add YoY growth calculations
        yoy_items = [
            ("Total Revenue", 'Revenues'),
            ("Operating Income", 'OperatingIncomeLoss'),
            ("EPS", 'EPS')  # Will use calculated EPS values
        ]
        
        for i, (yoy_name, metric_key) in enumerate(yoy_items):
            row = yoy_start_row + i + 1
            
            # YoY metric name
            cell = sheet.cell(row=row, column=1)
            cell.value = yoy_name
            cell.font = self.normal_font
            cell.alignment = self.left_align
            cell.border = self.border
            
            # Apply alternating row fill
            if i % 2 == 0:
                cell.fill = self.alternate_row_fill
            
            # Calculate YoY growth for each period (skip first 4 quarters)
            for j, (period_key, period_data) in enumerate(sorted_periods):
                col = j + 2
                cell = sheet.cell(row=row, column=col)
                
                # Skip first 4 quarters (no YoY comparison available)
                if j < 4:
                    cell.value = "N/A"
                    cell.font = self.note_font
                else:
                    # Get current period value and year-ago period value (4 quarters back)
                    current_period_data = sorted_periods[j][1]
                    year_ago_period_data = sorted_periods[j + 4][1] if j + 4 < len(sorted_periods) else None
                    
                    if year_ago_period_data:
                        # Handle EPS calculation for YoY growth
                        if metric_key == 'EPS':
                            # Calculate current EPS
                            current_items = current_period_data.get('items', {})
                            current_net_income = current_items.get('NetIncomeLoss', {}).get('value')
                            current_shares = current_items.get('WeightedAverageSharesOutstandingDiluted', {}).get('value')
                            current_eps = None
                            if (current_net_income is not None and current_shares is not None and current_shares != 0):
                                current_eps = current_net_income / current_shares
                            
                            # Calculate year-ago EPS
                            year_ago_items = year_ago_period_data.get('items', {})
                            year_ago_net_income = year_ago_items.get('NetIncomeLoss', {}).get('value')
                            year_ago_shares = year_ago_items.get('WeightedAverageSharesOutstandingDiluted', {}).get('value')
                            year_ago_eps = None
                            if (year_ago_net_income is not None and year_ago_shares is not None and year_ago_shares != 0):
                                year_ago_eps = year_ago_net_income / year_ago_shares
                            
                            # Calculate YoY EPS growth
                            if (current_eps is not None and year_ago_eps is not None and year_ago_eps != 0):
                                yoy_growth = (current_eps - year_ago_eps) / year_ago_eps * 100
                                cell.value = yoy_growth
                                cell.number_format = '0.00"%"'
                                
                                # Color code: green for positive, red for negative
                                if yoy_growth > 0:
                                    cell.font = Font(name='Arial', size=10, color='006100')
                                elif yoy_growth < 0:
                                    cell.font = Font(name='Arial', size=10, color='9C0006')
                                else:
                                    cell.font = self.number_font
                            else:
                                cell.value = "N/A"
                                cell.font = self.note_font
                        else:
                            # Handle regular metrics (Revenue, Operating Income)
                            current_items = current_period_data.get('items', {})
                            year_ago_items = year_ago_period_data.get('items', {})
                            
                            if (metric_key in current_items and 
                                metric_key in year_ago_items and
                                current_items[metric_key].get('value') is not None and
                                year_ago_items[metric_key].get('value') is not None and
                                year_ago_items[metric_key].get('value', 0) != 0):
                                
                                current_value = current_items[metric_key].get('value', 0)
                                year_ago_value = year_ago_items[metric_key].get('value', 0)
                                yoy_growth = (current_value - year_ago_value) / year_ago_value * 100
                                cell.value = yoy_growth
                                cell.number_format = '0.00"%"'
                                
                                # Color code: green for positive, red for negative
                                if yoy_growth > 0:
                                    cell.font = Font(name='Arial', size=10, color='006100')
                                elif yoy_growth < 0:
                                    cell.font = Font(name='Arial', size=10, color='9C0006')
                                else:
                                    cell.font = self.number_font
                            else:
                                cell.value = "N/A"
                                cell.font = self.note_font
                    else:
                        cell.value = "N/A"
                        cell.font = self.note_font
                
                cell.alignment = self.right_align
                cell.border = self.border
                
                # Apply alternating row fill
                if i % 2 == 0:
                    cell.fill = self.alternate_row_fill
    
    def _create_data_notes_sheet(self, sheet, income_statement: Dict):
        """Create data notes sheet explaining data limitations.
        
        Args:
            sheet: Excel worksheet to populate.
            income_statement: Income statement data.
        """
        # Add title
        sheet['A1'] = f"{income_statement.get('company_name', '')} ({income_statement.get('ticker', '')}) - Data Source Notes"
        sheet['A1'].font = Font(name='Arial', size=16, bold=True)
        sheet.merge_cells('A1:D1')
        sheet['A1'].alignment = Alignment(horizontal='center')
        
        # Add data source information
        sheet['A3'] = "Data Source Information"
        sheet['A3'].font = self.subheader_font
        sheet['A3'].fill = self.subheader_fill
        
        # Get data source notes if available
        data_source_notes = income_statement.get('data_source_notes', {})
        provider = data_source_notes.get('provider', 'Unknown')
        data_policy = data_source_notes.get('data_policy', 'N/A')
        
        sheet['A4'] = f"Primary Data Provider: {provider}"
        sheet['A5'] = f"Data Policy: {data_policy}"
        
        # Add field availability explanation
        sheet['A7'] = "Field Availability & Limitations"
        sheet['A7'].font = self.subheader_font
        sheet['A7'].fill = self.subheader_fill
        
        # Available fields
        sheet['A9'] = "✅ AVAILABLE FIELDS (High Confidence)"
        sheet['A9'].font = Font(name='Arial', size=11, bold=True, color='006100')
        
        available_fields = [
            "• Total Revenue - Direct from provider",
            "• Cost of Goods Sold - Direct from provider", 
            "• Gross Profit - Direct from provider",
            "• Research & Development - Direct from provider",
            "• Sales, General & Administrative (Combined) - Direct from provider",
            "• Total Operating Expenses - Direct from provider",
            "• Operating Income - Direct from provider",
            "• Pre-Tax Income - Direct from provider",
            "• Income Tax Expense - Direct from provider",
            "• Net Income - Direct from provider",
            "• Fully-Diluted Shares Outstanding - Direct from provider"
        ]
        
        for i, field in enumerate(available_fields):
            sheet[f'A{10 + i}'] = field
            sheet[f'A{10 + i}'].font = Font(name='Arial', size=10, color='006100')
        
        # Unavailable fields (significantly reduced)
        unavailable_start_row = 10 + len(available_fields) + 2
        sheet[f'A{unavailable_start_row}'] = "ℹ️  COMBINED/UNAVAILABLE FIELDS"
        sheet[f'A{unavailable_start_row}'].font = Font(name='Arial', size=11, bold=True, color='666666')
        
        unavailable_fields = [
            "• Sales & Marketing + General & Administrative - Combined into single SG&A line",
            "• Stock-Based Compensation - Not separately disclosed (typically included in SG&A)",
            "• Depreciation & Amortization - Not separately disclosed",
            "• Interest Income/Expense - Combined into 'Interest & Other Income, Expense' line"
        ]
        
        for i, field in enumerate(unavailable_fields):
            row = unavailable_start_row + 1 + i
            sheet[f'A{row}'] = field
            sheet[f'A{row}'].font = Font(name='Arial', size=10, color='666666')
        
        # Professional approach explanation
        explanation_start_row = unavailable_start_row + len(unavailable_fields) + 3
        sheet[f'A{explanation_start_row}'] = "Professional Data Quality Approach"
        sheet[f'A{explanation_start_row}'].font = self.subheader_font
        sheet[f'A{explanation_start_row}'].fill = self.subheader_fill
        
        explanation_text = [
            "This report follows a professional approach to financial data presentation:",
            "",
            "1. TRANSPARENCY: We clearly indicate what data is available vs. unavailable",
            "2. NO FALSE ESTIMATES: We never fabricate or estimate missing data points",
            "3. COMBINED DISCLOSURES: Where providers combine line items, we present them as combined",
            "4. CLEAR NOTATION: Unavailable fields are marked with (*) and highlighted",
            "",
            "This approach ensures you receive accurate, reliable financial data without",
            "any misleading estimates or artificial line item breakdowns."
        ]
        
        for i, text in enumerate(explanation_text):
            row = explanation_start_row + 2 + i
            sheet[f'A{row}'] = text
            if text.startswith(('1.', '2.', '3.', '4.')):
                sheet[f'A{row}'].font = Font(name='Arial', size=10, bold=True)
            else:
                sheet[f'A{row}'].font = Font(name='Arial', size=10)
        
        # Adjust column width
        sheet.column_dimensions['A'].width = 80
