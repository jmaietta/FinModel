"""
Vercel API function for generating financial reports.
Returns Excel data as base64 for client-side download.
"""
import os
import json
import base64
import io
import logging
from datetime import datetime
from flask import Flask, request, jsonify
import requests
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = Flask(__name__)

class PolygonDataFetcher:
    """Simplified Polygon data fetcher for Vercel."""
    
    def __init__(self, api_key):
        self.api_key = api_key
        self.base_url = "https://api.polygon.io"
    
    def get_income_statement(self, ticker, limit=12):
        """Fetch income statement data from Polygon."""
        try:
            endpoint = f"{self.base_url}/vX/reference/financials"
            params = {
                'ticker': ticker.upper(),
                'timeframe': 'quarterly',
                'limit': limit,
                'apikey': self.api_key
            }
            
            response = requests.get(endpoint, params=params, timeout=30)
            response.raise_for_status()
            data = response.json()
            
            if not data or 'results' not in data:
                return None
            
            # Process data into our format
            result = {
                'ticker': ticker.upper(),
                'company_name': ticker.upper(),
                'periods': {}
            }
            
            for api_result in data.get('results', []):
                if 'financials' in api_result and 'income_statement' in api_result['financials']:
                    income_stmt = api_result['financials']['income_statement']
                    period_date = api_result.get('end_date', '')
                    
                    if not period_date:
                        continue
                    
                    result['periods'][period_date] = {
                        'items': {
                            'Revenues': {
                                'value': self._safe_get_value(income_stmt, 'revenues'),
                                'source': 'polygon_direct'
                            },
                            'CostOfGoodsSold': {
                                'value': self._safe_get_value(income_stmt, 'cost_of_revenue'),
                                'source': 'polygon_direct'
                            },
                            'GrossProfit': {
                                'value': self._safe_get_value(income_stmt, 'gross_profit'),
                                'source': 'polygon_direct'
                            },
                            'ResearchAndDevelopmentExpense': {
                                'value': self._safe_get_value(income_stmt, 'research_and_development'),
                                'source': 'polygon_direct'
                            },
                            'OperatingExpenses': {
                                'value': self._safe_get_value(income_stmt, 'operating_expenses'),
                                'source': 'polygon_direct'
                            },
                            'OperatingIncomeLoss': {
                                'value': self._safe_get_value(income_stmt, 'operating_income_loss'),
                                'source': 'polygon_direct'
                            },
                            'NetIncomeLoss': {
                                'value': self._safe_get_value(income_stmt, 'net_income_loss'),
                                'source': 'polygon_direct'
                            }
                        }
                    }
            
            return result
            
        except Exception as e:
            logger.error(f"Error fetching data for {ticker}: {str(e)}")
            return None
    
    def _safe_get_value(self, data_dict, key, default=None):
        """Safely extract value from nested dictionary."""
        try:
            if key in data_dict and isinstance(data_dict[key], dict):
                value = data_dict[key].get('value')
                return float(value) if value is not None else None
            return default
        except (ValueError, TypeError):
            return default

class ExcelGenerator:
    """Generate Excel files in memory for Vercel."""
    
    def __init__(self):
        self.header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        self.normal_font = Font(name='Arial', size=10)
        self.header_fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type='solid')
        self.center_align = Alignment(horizontal='center', vertical='center')
        self.right_align = Alignment(horizontal='right', vertical='center')
        self.left_align = Alignment(horizontal='left', vertical='center')
        self.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
    
    def create_excel(self, income_statement):
        """Create Excel file in memory and return as bytes."""
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "Income Statement"
        
        # Add title
        sheet['A1'] = f"{income_statement.get('company_name', '')} ({income_statement.get('ticker', '')}) - Income Statement"
        sheet['A1'].font = Font(name='Arial', size=16, bold=True)
        sheet.merge_cells('A1:N1')
        sheet['A1'].alignment = Alignment(horizontal='center')
        
        # Extract and sort periods
        periods = income_statement.get('periods', {})
        sorted_periods = sorted(periods.items(), key=lambda x: x[0], reverse=True)[:12]
        
        if not sorted_periods:
            sheet['A3'] = "No data available"
            return self._workbook_to_bytes(wb)
        
        # Headers
        sheet['A3'] = "Line Item"
        sheet['A3'].font = self.header_font
        sheet['A3'].fill = self.header_fill
        sheet['A3'].alignment = self.center_align
        sheet['A3'].border = self.border
        
        # Period headers
        for i, (period_key, _) in enumerate(sorted_periods):
            col = i + 2
            cell = sheet.cell(row=3, column=col)
            cell.value = period_key
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_align
            cell.border = self.border
        
        # Line items
        line_items = [
            ('Total Revenue', 'Revenues'),
            ('Cost of Goods Sold', 'CostOfGoodsSold'),
            ('Gross Profit', 'GrossProfit'),
            ('Research & Development', 'ResearchAndDevelopmentExpense'),
            ('Operating Expenses', 'OperatingExpenses'),
            ('Operating Income', 'OperatingIncomeLoss'),
            ('Net Income', 'NetIncomeLoss')
        ]
        
        for i, (display_name, item_key) in enumerate(line_items):
            row = i + 4
            
            # Item name
            cell = sheet.cell(row=row, column=1)
            cell.value = display_name
            cell.font = self.normal_font
            cell.alignment = self.left_align
            cell.border = self.border
            
            # Values for each period
            for j, (period_key, period_data) in enumerate(sorted_periods):
                col = j + 2
                cell = sheet.cell(row=row, column=col)
                
                items = period_data.get('items', {})
                if item_key in items and items[item_key].get('value') is not None:
                    value = items[item_key]['value']
                    cell.value = value
                    cell.number_format = '$#,##0,,"M"'  # Display in millions
                else:
                    cell.value = "N/A"
                
                cell.font = self.normal_font
                cell.alignment = self.right_align
                cell.border = self.border
        
        # Adjust column widths
        sheet.column_dimensions['A'].width = 25
        for col in range(2, len(sorted_periods) + 2):
            sheet.column_dimensions[get_column_letter(col)].width = 15
        
        return self._workbook_to_bytes(wb)
    
    def _workbook_to_bytes(self, workbook):
        """Convert workbook to bytes."""
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        return output.getvalue()

@app.route('/', methods=['GET'])
def home():
    """Serve the main page."""
    return '''<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1.0"/>
  <title>FinModel - Financial Statement Generator</title>
  <link 
    href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" 
    rel="stylesheet"
  />
  <style>
    body {
      font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
      background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
      min-height: 100vh;
      color: #212529;
      margin: 0; 
      padding: 0;
    }
    .main-container {
      min-height: 100vh;
      display: flex;
      align-items: center;
      justify-content: center;
      padding: 20px;
    }
    .card {
      border-radius: 15px;
      box-shadow: 0 15px 35px rgba(0,0,0,0.1);
      backdrop-filter: blur(10px);
      background: rgba(255,255,255,0.95);
      border: 1px solid rgba(255,255,255,0.2);
      max-width: 500px;
      width: 100%;
    }
    .card-header {
      background: linear-gradient(135deg, #0066cc 0%, #004499 100%);
      color: white;
      font-weight: bold;
      border-radius: 15px 15px 0 0;
      text-align: center;
      padding: 20px;
    }
    .card-body {
      padding: 30px;
    }
    .btn-primary {
      background: linear-gradient(135deg, #0066cc 0%, #004499 100%);
      border: none;
      border-radius: 8px;
      padding: 12px 30px;
      font-weight: 600;
      transition: all 0.3s ease;
    }
    .btn-primary:hover {
      transform: translateY(-2px);
      box-shadow: 0 5px 15px rgba(0,102,204,0.4);
    }
    .btn-success {
      background: linear-gradient(135deg, #28a745 0%, #20c997 100%);
      border: none;
      border-radius: 8px;
      padding: 12px 30px;
      font-weight: 600;
    }
    .ticker-input {
      font-size: 1.2rem;
      text-transform: uppercase;
      border-radius: 8px;
      border: 2px solid #e9ecef;
      transition: border-color 0.3s ease;
    }
    .ticker-input:focus {
      border-color: #0066cc;
      box-shadow: 0 0 0 0.2rem rgba(0,102,204,0.25);
    }
    .loading {
      display: none;
      text-align: center;
      margin: 20px 0;
    }
    .loading .spinner-border {
      width: 3rem;
      height: 3rem;
    }
    .result-section {
      display: none;
      margin-top: 20px;
    }
    .brand-title {
      font-size: 2.5rem;
      font-weight: 700;
      background: linear-gradient(135deg, #0066cc 0%, #764ba2 100%);
      -webkit-background-clip: text;
      -webkit-text-fill-color: transparent;
      background-clip: text;
      margin-bottom: 10px;
    }
    .subtitle {
      color: #6c757d;
      font-size: 1.1rem;
      margin-bottom: 30px;
    }
    .feature-list {
      list-style: none;
      padding: 0;
      margin: 20px 0;
    }
    .feature-list li {
      padding: 8px 0;
      color: #495057;
    }
    .feature-list li:before {
      content: "✓";
      color: #28a745;
      font-weight: bold;
      margin-right: 10px;
    }
  </style>
</head>
<body>
  <div class="main-container">
    <div class="card">
      <div class="card-header">
        <h1 class="brand-title mb-0">FinModel</h1>
        <div class="subtitle mb-0">Professional Financial Analysis</div>
      </div>
      
      <div class="card-body">
        <form id="ticker-form">
          <div class="mb-4">
            <label for="ticker" class="form-label fw-bold">Stock Ticker Symbol</label>
            <input
              type="text"
              class="form-control ticker-input"
              id="ticker"
              placeholder="e.g., AAPL, MSFT, GOOGL"
              required
            />
            <div class="form-text">
              Enter any publicly-traded technology company ticker.
            </div>
          </div>
          
          <ul class="feature-list">
            <li>12 quarters of historical data</li>
            <li>Professional Excel format</li>
            <li>Real-time financial data</li>
            <li>Institutional-grade analysis</li>
          </ul>
          
          <button type="submit" class="btn btn-primary w-100">
            Generate Income Statement
          </button>
        </form>

        <div class="loading" id="loading-section">
          <div class="spinner-border text-primary" role="status">
            <span class="visually-hidden">Loading...</span>
          </div>
          <p class="mt-3 fw-bold">
            Retrieving financial data...
          </p>
          <p class="text-muted small">
            Analyzing multiple quarters of data to build your model.
          </p>
        </div>

        <div class="result-section" id="result-section">
          <div class="alert alert-success">
            <h5 class="alert-heading">✓ Income Statement Ready</h5>
            <p class="mb-0">Your comprehensive financial model has been generated successfully.</p>
          </div>
          
          <div class="d-grid gap-2">
            <button class="btn btn-success" id="download-btn">
              📊 Download Excel File
            </button>
            <button class="btn btn-outline-secondary" id="new-search">
              🔄 Generate Another Report
            </button>
          </div>
        </div>
      </div>
    </div>
  </div>

  <script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/js/bootstrap.bundle.min.js"></script>
  <script>
    document.addEventListener('DOMContentLoaded', function() {
      const tickerForm = document.getElementById('ticker-form');
      const loadingSection = document.getElementById('loading-section');
      const resultSection = document.getElementById('result-section');
      const downloadBtn = document.getElementById('download-btn');
      const newSearchBtn = document.getElementById('new-search');
      
      let excelData = null;
      let filename = null;

      tickerForm.addEventListener('submit', function(e) {
        e.preventDefault();
        const ticker = document.getElementById('ticker').value.trim().toUpperCase();
        if (!ticker) return;

        // Show loading
        loadingSection.style.display = 'block';
        resultSection.style.display = 'none';

        fetch('/api/generate', {
          method: 'POST',
          headers: { 'Content-Type': 'application/json' },
          body: JSON.stringify({ ticker })
        })
        .then(response => response.json())
        .then(data => {
          loadingSection.style.display = 'none';
          
          if (data.success) {
            excelData = data.excel_data;
            filename = data.filename;
            resultSection.style.display = 'block';
          } else {
            alert('Error: ' + (data.error || 'Failed to generate report'));
          }
        })
        .catch(err => {
          loadingSection.style.display = 'none';
          alert('Error: ' + err.message);
        });
      });

      downloadBtn.addEventListener('click', function() {
        if (excelData && filename) {
          // Convert base64 to blob and download
          const byteCharacters = atob(excelData);
          const byteNumbers = new Array(byteCharacters.length);
          for (let i = 0; i < byteCharacters.length; i++) {
            byteNumbers[i] = byteCharacters.charCodeAt(i);
          }
          const byteArray = new Uint8Array(byteNumbers);
          const blob = new Blob([byteArray], { 
            type: 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' 
          });
          
          const url = window.URL.createObjectURL(blob);
          const a = document.createElement('a');
          a.href = url;
          a.download = filename;
          document.body.appendChild(a);
          a.click();
          window.URL.revokeObjectURL(url);
          document.body.removeChild(a);
        }
      });

      newSearchBtn.addEventListener('click', function() {
        resultSection.style.display = 'none';
        document.getElementById('ticker').value = '';
        excelData = null;
        filename = null;
      });
    });
  </script>
</body>
</html>'''

@app.route('/api/generate', methods=['POST'])
def generate():
    """Generate financial report API endpoint."""
    try:
        # Get request data
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'error': 'No data provided'}), 400
        
        ticker = data.get('ticker', '').strip().upper()
        if not ticker:
            return jsonify({'success': False, 'error': 'Ticker symbol is required'}), 400
        
        # Get API key from environment
        api_key = os.environ.get('POLYGON_API_KEY')
        if not api_key:
            return jsonify({'success': False, 'error': 'API key not configured'}), 500
        
        # Fetch data
        fetcher = PolygonDataFetcher(api_key)
        income_statement = fetcher.get_income_statement(ticker)
        
        if not income_statement or not income_statement.get('periods'):
            return jsonify({'success': False, 'error': f'No data found for {ticker}'}), 404
        
        # Generate Excel
        generator = ExcelGenerator()
        excel_bytes = generator.create_excel(income_statement)
        
        # Convert to base64 for JSON response
        excel_b64 = base64.b64encode(excel_bytes).decode('utf-8')
        
        return jsonify({
            'success': True,
            'ticker': ticker,
            'company_name': income_statement.get('company_name', ticker),
            'excel_data': excel_b64,
            'filename': f"{ticker}_Income_Statement.xlsx"
        })
        
    except Exception as e:
        logger.error(f"Error in generate endpoint: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

# Vercel expects this format
def handler(request):
    return app(request.environ, lambda *args: None)
